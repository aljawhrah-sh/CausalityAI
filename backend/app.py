from flask import Flask, send_from_directory, request, jsonify
import os
import database

app = Flask(__name__)

SCREENS = os.path.join(os.path.dirname(__file__), '..', 'screens')

def build_breakdown(case):
    dechallenge = (case.get('dechallenge') or '').lower()
    rechallenge = (case.get('rechallenge') or '').lower()
    criteria = []

    if dechallenge == 'positive':
        criteria.append({'label': 'Positive dechallenge', 'detail': 'Symptoms resolved after withdrawal', 'weight': 3})
    elif dechallenge == 'negative':
        criteria.append({'label': 'Negative dechallenge', 'detail': 'Symptoms persisted after withdrawal', 'weight': -1})

    if rechallenge == 'positive':
        criteria.append({'label': 'Positive rechallenge', 'detail': 'Reaction recurred on re-exposure', 'weight': 3})

    if case.get('time_onset'):
        criteria.append({'label': 'Time to onset recorded', 'detail': case['time_onset'], 'weight': 1})

    if case.get('narrative'):
        criteria.append({'label': 'Clinical narrative provided', 'detail': 'Free-text description available', 'weight': 1})

    if case.get('alternative'):
        criteria.append({'label':'Alternative cause present', 'detail': case['alternative'], 'weight': -1})

    return criteria

#dechallenge scoring method
def score_dechallenge(performed, resolved):
    p = (performed or '').strip().lower()
    r = (resolved or '').strip().lower()

    if p == 'yes' and r == 'yes': return 3
    if p == 'yes' and r == 'no': return -1
    if p == 'yes': return 1
    if p == 'no' and r == 'yes': return -1 #we will need SFDA confirmation
    return 0

#routing/mapping homepage url
@app.route('/')
def home():
    return send_from_directory(SCREENS, '00-login.html')

#filename ->whatever html inside screens gets in the function 
#http://127.0.0.1:5000/screens/any html file
@app.route('/screens/<filename>')
def serve_screen(filename):
    return send_from_directory(SCREENS, filename)

@app.route('/api/assess',methods=['POST'])
def assess():
    # we'll be using numbers as wieghts for scoring ex:score >= 6 -> certain (92%)
    #later on we will change these placeholders +1,-1, +3 into real patterns for the ML model

    data = request.get_json()

    dechallenge = data.get('dechallenge', 'unknown')
    dechallenge_resolved = data.get('dechallenge_resolved', '')
    rechallenge = data.get('rechallenge', 'unknown')
    time_to_onset = data.get('time_to_onset', '')
    alternative = data.get('alternative_cause', '')
    narrative = data.get('narrative', '')
    age = data.get('age', '')
    sex = data.get('sex', '')
    region = data.get('region', '')
    reporter = data.get('reporter', '')

    
    #check if dechallenge has a real value 
    #check if has_evidence is true
    has_dechallenge = (dechallenge or '').strip().lower() in ('yes', 'no')
    has_rechallenge = rechallenge in ('positive', 'negative')
    has_evidence = has_dechallenge or has_rechallenge or time_to_onset or narrative

    score = 0
    score += score_dechallenge(dechallenge, dechallenge_resolved)
    if rechallenge == 'positive':
        score += 3
    if time_to_onset:
        score +=1
    if alternative:
        score -= 1
    if narrative:
        score += 1

    if not has_evidence:
        category = 'Unassessable'
        confidence = 20
    elif score >= 6:
        category = 'Certain'
        confidence = 92
    elif score >= 4:
        category = 'Probable / Likely'
        confidence = 78
    elif score >= 2:
        category = 'Possible'
        confidence = 61
    elif score >= 0:
        category = 'Unlikely'
        confidence = 45
    else:
        category = 'Unassessable'
        confidence = 20

    case_id = database.save_case(
        drug = data.get('drug_name', ''),
        age = age,
        sex = sex,
        region = region,
        time_onset = time_to_onset,
        dechallenge = dechallenge,
        dechallenge_resolved = dechallenge_resolved,
        rechallenge = rechallenge,
        alternative = alternative,
        narrative = narrative,
        category = category,
        confidence = confidence,
        score = score
    )

    return jsonify({
        'case_id': case_id,
        'category': category,
        'confidence' : confidence,
        'score' : score,
        'patient':{
            'age': age,
            'sex': sex,
            'region': region,
            'reporter': reporter
        },
        'drug': data.get('drug_name', ''),
        'time_to_onset': data.get('time_to_onset', '')
    })

@app.route('/api/decision', methods=['POST'])
def decision():
    data = request.get_json()

    case_id = data.get('case_id', 0)
    assessor = data.get('assessor', 'N A')
    action = data.get('action', '')
    final_cat = data.get('final_cat', '')
    reasoning = data.get('reasoning', '')

    database.save_decision(case_id, assessor, action, final_cat, reasoning)

    return jsonify({
        'success': True,
        'message': 'Decision logged successfully'
    })

@app.route('/api/cases', methods=['GET'])
def get_cases():
    cases = database.get_cases()
    return jsonify(cases)

@app.route('/api/case/<int:case_id>', methods=['GET'])
def get_case(case_id):
    case = database.get_case_by_id(case_id)

    if case is None:
        return jsonify({'success': False, 'message':'Case not found'}), 404
    case['breakdown'] = build_breakdown(case)
    return jsonify(case)

@app.route('/api/upload', methods=['POST'])
def upload_excel():
    #two ifs checking if the file was sent before trying to read it  
    if 'file' not in request.files:
        return jsonify({'success': False, 'message':'No file uploaded'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'message': 'No files selected'}), 400
    
    try:
        import openpyxl
        #wb = woekbook which is the whole excel file
        wb = openpyxl.load_workbook(file)
        #ws = worksheet which is the first sheet in the workbook
        ws = wb.active

        headers = []
        #ws[1] is the header that's why we skipped the first row the the secnd will be the data and so on
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append('')
        #counter
        cases_processed = 0
        all_cases = []

        for row in ws.iter_rows(min_row= 2, values_only= True):
            #if the entire row is empty jump to the next one 
            if not any(row):
                continue
            #ex: {'drug_name': 'Warfarin 5mg', 'age': 45, 'sex': 'F'}
            data = dict(zip(headers, row))

            drug = str(data.get('drug_name', '') or '').strip()
            age = str(data.get('age', '') or '').strip()
            sex = str(data.get('sex', '')or '').strip()
            region = str(data.get('region', '')or '').strip()
            time_onset = str(data.get('time_to_onset', '') or '').strip()
            dechallenge = str(data.get('dechallenge', 'unknown') or 'unknown').strip().lower()
            dechallenge_resolved = str(data.get('dechallenge_resolved', '') or '').strip().lower()
            rechallenge = str(data.get('rechallenge', 'unknown') or 'unknown').strip().lower()
            alternative = str(data.get('alternalive_cause', '') or '').strip()
            narrative = str(data.get('narrative', '') or '').strip()


            #check if dechallenge has a real value 
            #check if has_evidence is true
            has_dechallenge = dechallenge in ('yes', 'no')
            has_rechallenge = rechallenge in ('positive', 'negative')
            has_evidence = has_dechallenge or has_rechallenge or time_onset or narrative

            #scoring logic same as assess() applied to each row
            score = 0
            score += score_dechallenge(dechallenge, dechallenge_resolved)
            if rechallenge == 'positive':
                score += 3
            if time_onset:
                score += 1
            if alternative:
                score -= 1
            if narrative:
                score += 1


            # mapping by converting numeric score into one of our six WHO-WMC categories with its percentage
            if not has_evidence:
                category = 'Unassessable'
                confidence = 20
            elif score >= 6:
                category = 'Certain'
                confidence = 92
            elif score >= 4:
                category = 'Probable / Likely'
                confidence = 78
            elif score >= 2:
                category = 'Possible'
                confidence = 61
            elif score >= 0:
                category = 'Unlikely'
                confidence = 45
            else:
                category = 'Enassessable'
                confidence = 20

            #save to database
            case_id = database.save_case(
                drug = drug,
                age = age,
                sex = sex,
                region = region,
                time_onset= time_onset,
                dechallenge= dechallenge,
                dechallenge_resolved = dechallenge_resolved,
                rechallenge = rechallenge,
                alternative = alternative,
                narrative= narrative,
                category= category,
                confidence= confidence,
                score= score
            )

            #sending the case back to frontend for the preview
            all_cases.append({
                'id': case_id,
                'drug': drug,
                'age': age,
                'sex': sex,
                'region': region,
                'time_onset': time_onset,
                'dechallenge': dechallenge,
                'rechallenge': rechallenge,
                'alternative': alternative,
                'category': category,
                'confidence': confidence,
                'score': score
            })

            cases_processed += 1
        return jsonify({
            'success': True,
            'cases_processed': cases_processed,
            'cases': all_cases,
            'message': f'{cases_processed} cases assessed and saved successfully'
        })    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/api/stats', methods=['GET'])
def get_stats():
    import sqlite3
    conn = sqlite3.connect(database.DB_PATH)
    cursor = conn.cursor()

    #total cases in queue
    cursor.execute('SELECT COUNT(*) FROM cases')
    total_cases = cursor.fetchone()[0]

    #Cases that are completed today
    cursor.execute(''' SELECT COUNT(*) FROM decisions WHERE DATE(decided_at) = DATE('now')''')
    completed_today = cursor.fetchone()[0]

    #confirmed cases vs overriden cases
    cursor.execute("SELECT COUNT(*) FROM decisions WHERE action = 'confirmed'")
    confirmed = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(*) FROM decisions')
    total_decisions = cursor.fetchone()[0]

    #counts for category
    cursor.execute(''' SELECT category, COUNT(*) FROM cases GROUP BY category''')
    category_rows = cursor.fetchall()
    categories = {}
    for row in category_rows:
        categories[row[0]] = row[1]

    conn.close()

    #agreement rounding calculation
    agreement = round((confirmed / total_decisions * 100)) if total_decisions > 0 else 0

    return jsonify({
        'total_cases': total_cases,
        'completed_today': completed_today,
        'agreement_rate': agreement,
        'total_decisions':total_decisions,
        'categories': categories
    })

#debug true -> server restart automatically, port 5000 -> run http://127.0.0.1:5000
if __name__ == '__main__':
    database.create_tables()
    app.run(debug=True, port=5000)

